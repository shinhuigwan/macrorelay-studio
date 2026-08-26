import argparse
import json
import io
import sys
import os
import shutil
import re
import time
from datetime import datetime
from pathlib import Path

try:
    from PIL import ImageGrab, Image
except Exception as exc:  # pragma: no cover
    raise RuntimeError("Pillow가 설치되어 있지 않습니다. pip install pillow") from exc

try:
    import pytesseract
except Exception as exc:  # pragma: no cover
    raise RuntimeError("pytesseract가 설치되어 있지 않습니다. pip install pytesseract") from exc


def find_tesseract_path() -> str | None:
    helper_dir = Path(__file__).resolve().parent
    configured_files = [helper_dir / "tesseract_path.txt", helper_dir.parent / "tesseract_path.txt"]
    for configured_file in configured_files:
        try:
            raw_path = configured_file.read_text(encoding="utf-8-sig").strip()
        except (OSError, ValueError):
            continue
        if not raw_path:
            continue
        configured = Path(raw_path)
        if not configured.is_absolute():
            configured = helper_dir / configured
        if configured.exists():
            return str(configured)
    candidates = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return candidate
    return shutil.which("tesseract")


def _resolve_tessdata_dir(tesseract_path: Path) -> Path | None:
    base = tesseract_path.parent
    candidate = base / "tessdata"
    if candidate.exists():
        return candidate
    return None


def _resolve_tessdata_dir_env() -> Path | None:
    env_dir = os.environ.get("TESSDATA_PREFIX")
    if env_dir:
        candidate = Path(env_dir)
        if candidate.exists():
            return candidate
    return None


def ensure_tesseract(lang: str = "eng+kor", tessdata_variant: str = "fast") -> None:
    cmd = getattr(pytesseract.pytesseract, "tesseract_cmd", "") or ""
    if not cmd or not Path(cmd).exists():
        found = find_tesseract_path()
        if not found:
            raise RuntimeError("Tesseract OCR is not installed or not on PATH.")
        pytesseract.pytesseract.tesseract_cmd = found
        cmd = found

    wanted = [code.strip() for code in (lang or "").split("+") if code.strip()]
    candidates = []
    env_dir = os.environ.get("TESSDATA_PREFIX")
    if env_dir:
        candidates.append(Path(env_dir))
    install_dir = _resolve_tessdata_dir(Path(cmd))
    if install_dir:
        candidates.append(install_dir)
    local_dir = Path(__file__).resolve().parent / ("tessdata_best" if tessdata_variant == "best" else "tessdata")
    if local_dir.exists():
        candidates.append(local_dir)
    alt_local = Path(__file__).resolve().parent / "tessdata"
    if alt_local.exists() and alt_local not in candidates:
        candidates.append(alt_local)

    def has_all(folder: Path) -> bool:
        return all((folder / f"{code}.traineddata").exists() for code in wanted)

    for folder in candidates:
        if folder.exists() and has_all(folder):
            os.environ["TESSDATA_PREFIX"] = str(folder)
            return

    # Merge missing langs into local tessdata folder if possible.
    if not local_dir.exists():
        local_dir.mkdir(parents=True, exist_ok=True)
    for code in wanted:
        target = local_dir / f"{code}.traineddata"
        if target.exists():
            continue
        for folder in candidates:
            src = folder / f"{code}.traineddata"
            if src.exists():
                shutil.copy2(src, target)
                break

    if local_dir.exists() and has_all(local_dir):
        os.environ["TESSDATA_PREFIX"] = str(local_dir)
        return

    missing_all = ", ".join([code for code in wanted if not (local_dir / f"{code}.traineddata").exists()])
    raise RuntimeError("Tesseract language data missing: " + missing_all)


def capture_region(left: int, top: int, right: int, bottom: int) -> Image.Image:
    bbox = (left, top, right, bottom)
    try:
        return ImageGrab.grab(bbox=bbox, all_screens=True)
    except TypeError:
        return ImageGrab.grab(bbox=bbox)


def capture_browser(left: int, top: int, right: int, bottom: int) -> Image.Image:
    bbox = (left, top, right, bottom)
    try:
        return ImageGrab.grab(bbox=bbox, all_screens=True)
    except TypeError:
        return ImageGrab.grab(bbox=bbox)


def find_element(page, selector: str, timeout_ms: int, poll_ms: int):
    end = time.time() + (timeout_ms / 1000.0)
    while time.time() < end:
        for frame in page.frames:
            try:
                element = frame.query_selector(selector)
            except Exception:
                element = None
            if element:
                return element
        time.sleep(poll_ms / 1000.0)
    return None


def capture_browser(selector: str, title: str, port: int) -> tuple[Image.Image, str]:
    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:
        raise RuntimeError("Playwright가 설치되어 있지 않습니다.") from exc

    url = f"http://127.0.0.1:{port}"
    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp(url)
        page = None
        if title:
            for context in browser.contexts:
                for pge in context.pages:
                    try:
                        if title in pge.title():
                            page = pge
                            break
                    except Exception:
                        continue
                if page:
                    break
        if page is None:
            for context in browser.contexts:
                if context.pages:
                    page = context.pages[0]
                    break
        if page is None:
            raise RuntimeError("대상 페이지를 찾지 못했습니다.")
        element = find_element(page, selector, 2000, 50)
        if not element:
            raise RuntimeError("요소를 찾지 못했습니다.")
        dom_text = ""
        try:
            dom_text = element.evaluate(
                """
                (el) => {
                  if (!el) return '';
                  const tag = (el.tagName || '').toLowerCase();
                  if (tag === 'input' || tag === 'textarea') {
                    return (el.value || '').trim();
                  }
                  const txt = (el.innerText || el.textContent || '').trim();
                  if (txt) return txt;
                  try {
                    const walker = document.createTreeWalker(el, NodeFilter.SHOW_TEXT, null);
                    let node = walker.nextNode();
                    while (node) {
                      const value = (node.nodeValue || '').trim();
                      if (value) return value;
                      node = walker.nextNode();
                    }
                  } catch (err) {}
                  try {
                    const before = window.getComputedStyle(el, '::before').content;
                    if (before && before !== 'none' && before !== 'normal') {
                      return String(before).replace(/^["']|["']$/g, '').trim();
                    }
                    const after = window.getComputedStyle(el, '::after').content;
                    if (after && after !== 'none' && after !== 'normal') {
                      return String(after).replace(/^["']|["']$/g, '').trim();
                    }
                  } catch (err) {}
                  const attrs = ['aria-label', 'title', 'alt', 'placeholder', 'value', 'data-tooltip'];
                  for (const name of attrs) {
                    const val = el.getAttribute && el.getAttribute(name);
                    if (val && String(val).trim()) return String(val).trim();
                  }
                  return '';
                }
                """
            ) or ""
            dom_text = dom_text.strip()
        except Exception:
            dom_text = ""
        img_bytes = element.screenshot(type="png")
        browser.close()
    return Image.open(io.BytesIO(img_bytes)), dom_text


def preprocess_image(image: Image.Image, scale: int = 2) -> Image.Image:
    # Basic preprocessing to improve OCR accuracy.
    gray = image.convert("L")
    w, h = gray.size
    scale = max(1, int(scale))
    resized = gray.resize((w * scale, h * scale), Image.BICUBIC)
    # Simple auto-threshold based on average luminance.
    pixels = resized.getdata()
    avg = sum(pixels) / max(1, len(pixels))
    threshold = max(80, min(200, int(avg)))
    bw = resized.point(lambda p: 255 if p > threshold else 0, mode="1")
    return bw


def preprocess_fallback(image: Image.Image, scale: int = 2) -> Image.Image:
    gray = image.convert("L")
    w, h = gray.size
    scale = max(1, int(scale))
    resized = gray.resize((w * scale, h * scale), Image.BICUBIC)
    return resized


def normalize_text(text: str) -> str:
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    cleaned = [line.rstrip() for line in lines if line.rstrip()]
    return "\n".join(cleaned)


def write_output(text: str, path: str, fmt: str, append: bool) -> None:
    if not path:
        return
    fmt = fmt.lower()
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    if fmt == "json":
        entry = {"timestamp": datetime.utcnow().isoformat() + "Z", "text": text}
        if target.exists() and append:
            try:
                data = json.loads(target.read_text(encoding="utf-8-sig"))
            except Exception:
                data = []
            if isinstance(data, list):
                data.append(entry)
                target.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8-sig")
                return
        target.write_text(json.dumps([entry], ensure_ascii=False, indent=2), encoding="utf-8-sig")
    elif fmt == "txt":
        mode = "a" if append else "w"
        with target.open(mode, encoding="utf-8-sig") as handle:
            handle.write(text + "\n")
    else:
        line = f"{datetime.utcnow().isoformat()}Z,{text}\n"
        mode = "a" if append else "w"
        with target.open(mode, encoding="utf-8-sig") as handle:
            handle.write(line)


def write_excel(text: str, mode: str, path: str, sheet: str, cell: str) -> None:
    mode = (mode or "none").lower()
    if mode == "none":
        return
    if mode == "file":
        try:
            import openpyxl
        except Exception as exc:
            raise RuntimeError("openpyxl이 설치되어 있지 않습니다.") from exc
        target = Path(path) if path else None
        if not target:
            raise RuntimeError("엑셀 파일 경로가 필요합니다.")
        if target.exists():
            wb = openpyxl.load_workbook(target)
        else:
            wb = openpyxl.Workbook()
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        ws[cell or "A1"] = text
        wb.save(target)
    elif mode == "running":
        try:
            import win32com.client
        except Exception as exc:
            raise RuntimeError("win32com이 설치되어 있지 않습니다.") from exc
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
        except Exception as exc:
            raise RuntimeError("실행 중인 엑셀을 찾지 못했습니다.") from exc
        if path:
            wb = None
            for wb_item in excel.Workbooks:
                if Path(wb_item.FullName).samefile(Path(path)):
                    wb = wb_item
                    break
            if wb is None:
                wb = excel.Workbooks.Open(path)
        else:
            wb = excel.ActiveWorkbook
        if wb is None:
            raise RuntimeError("엑셀 워크북이 없습니다.")
        ws = wb.Worksheets(sheet) if sheet else wb.ActiveSheet
        ws.Range(cell or "A1").Value = text
    else:
        raise RuntimeError(f"지원하지 않는 excel 모드: {mode}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", default="region")
    parser.add_argument("--left", type=int, default=0)
    parser.add_argument("--top", type=int, default=0)
    parser.add_argument("--right", type=int, default=0)
    parser.add_argument("--bottom", type=int, default=0)
    parser.add_argument("--selector", default="")
    parser.add_argument("--selector-file", default="")
    parser.add_argument("--title", default="")
    parser.add_argument("--title-file", default="")
    parser.add_argument("--port", type=int, default=9222)
    parser.add_argument("--lang", default="eng+kor")
    parser.add_argument("--psm", type=int, default=6)
    parser.add_argument("--scale", type=int, default=2)
    parser.add_argument("--whitelist", default="")
    parser.add_argument("--tessdata", default="fast")
    parser.add_argument("--stdout", action="store_true")
    parser.add_argument("--debug-out", default="")
    parser.add_argument("--output", default="")
    parser.add_argument("--output-format", default="csv")
    parser.add_argument("--append", default="1")
    parser.add_argument("--also-output", default="")
    parser.add_argument("--also-format", default="")
    parser.add_argument("--also-append", default="")
    parser.add_argument("--excel-mode", default="none")
    parser.add_argument("--excel-path", default="")
    parser.add_argument("--excel-sheet", default="")
    parser.add_argument("--excel-cell", default="")
    args = parser.parse_args()

    ensure_tesseract(args.lang, args.tessdata)

    mode = args.mode.lower()
    if mode == "browser":
        if not args.selector and args.selector_file:
            try:
                with open(args.selector_file, "r", encoding="utf-8-sig") as handle:
                    args.selector = handle.read().strip()
            except Exception:
                args.selector = ""
        if not args.title and args.title_file:
            try:
                with open(args.title_file, "r", encoding="utf-8-sig") as handle:
                    args.title = handle.read().strip()
            except Exception:
                args.title = ""
        if not args.selector:
            raise RuntimeError("selector가 필요합니다.")
        image, dom_text = capture_browser(args.selector, args.title, args.port)
        if dom_text:
            text = normalize_text(dom_text)
            write_output(text, args.output, args.output_format, args.append == "1")
            print(text)
            return
    else:
        if args.right <= args.left or args.bottom <= args.top:
            raise RuntimeError("좌표 범위가 올바르지 않습니다.")
        image = capture_region(args.left, args.top, args.right, args.bottom)

    debug_out = args.debug_out.strip()
    debug_base = Path(debug_out) if debug_out else None
    if debug_base:
        debug_base.parent.mkdir(parents=True, exist_ok=True)
        try:
            image.save(debug_base)
        except Exception:
            pass
    try:
        config = f"--psm {args.psm} --oem 3 -c preserve_interword_spaces=1"
        if args.whitelist:
            config += f" -c tessedit_char_whitelist={args.whitelist}"
        prepped = preprocess_image(image, args.scale)
        if debug_base:
            try:
                prepped.save(debug_base.with_name(debug_base.stem + "_pre.png"))
            except Exception:
                pass
        raw_text = pytesseract.image_to_string(prepped, lang=args.lang, config=config)
        text = normalize_text(raw_text)
        if not text:
            fallback = preprocess_fallback(image, args.scale)
            if debug_base:
                try:
                    fallback.save(debug_base.with_name(debug_base.stem + "_raw.png"))
                except Exception:
                    pass
            raw_text = pytesseract.image_to_string(fallback, lang=args.lang, config=config)
            text = normalize_text(raw_text)
        if not text:
            raw_text = pytesseract.image_to_string(image, lang=args.lang, config=config)
            text = normalize_text(raw_text)
    except pytesseract.TesseractNotFoundError:
        raise RuntimeError("Tesseract OCR is not installed or not on PATH.")
    write_output(text, args.output, args.output_format, args.append != "0")
    if args.also_output:
        alt_format = args.also_format or args.output_format
        alt_append = args.also_append if args.also_append != "" else args.append
        write_output(text, args.also_output, alt_format, str(alt_append) != "0")
    write_excel(text, args.excel_mode, args.excel_path, args.excel_sheet, args.excel_cell)
    print(text)


if __name__ == "__main__":
    main()



