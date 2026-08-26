import argparse
from pathlib import Path


def _read_excel_file(path: str, sheet: str, cell: str) -> str:
    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("openpyxl is not installed.") from exc
    target = Path(path)
    if not target.exists():
        raise RuntimeError("Excel file not found.")
    wb = openpyxl.load_workbook(target, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    value = ws[cell or "A1"].value
    return "" if value is None else str(value)


def _write_excel_file(path: str, sheet: str, cell: str, value: str) -> None:
    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("openpyxl is not installed.") from exc
    target = Path(path)
    if target.exists():
        wb = openpyxl.load_workbook(target)
    else:
        wb = openpyxl.Workbook()
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    ws[cell or "A1"] = value
    wb.save(target)


def _read_excel_running(path: str, sheet: str, cell: str) -> str:
    try:
        import win32com.client
    except Exception as exc:
        raise RuntimeError("pywin32 is not installed.") from exc
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
    except Exception as exc:
        raise RuntimeError("No running Excel instance.") from exc
    if path:
        wb = None
        for wb_item in excel.Workbooks:
            try:
                if Path(wb_item.FullName).samefile(Path(path)):
                    wb = wb_item
                    break
            except Exception:
                continue
        if wb is None:
            wb = excel.Workbooks.Open(path)
    else:
        wb = excel.ActiveWorkbook
    if wb is None:
        raise RuntimeError("Excel workbook not available.")
    ws = wb.Worksheets(sheet) if sheet else wb.ActiveSheet
    value = ws.Range(cell or "A1").Value
    return "" if value is None else str(value)


def _write_excel_running(path: str, sheet: str, cell: str, value: str) -> None:
    try:
        import win32com.client
    except Exception as exc:
        raise RuntimeError("pywin32 is not installed.") from exc
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
    except Exception as exc:
        raise RuntimeError("No running Excel instance.") from exc
    if path:
        wb = None
        for wb_item in excel.Workbooks:
            try:
                if Path(wb_item.FullName).samefile(Path(path)):
                    wb = wb_item
                    break
            except Exception:
                continue
        if wb is None:
            wb = excel.Workbooks.Open(path)
    else:
        wb = excel.ActiveWorkbook
    if wb is None:
        raise RuntimeError("Excel workbook not available.")
    ws = wb.Worksheets(sheet) if sheet else wb.ActiveSheet
    ws.Range(cell or "A1").Value = value


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", default="file")
    parser.add_argument("--path", default="")
    parser.add_argument("--sheet", default="")
    parser.add_argument("--cell", default="")
    parser.add_argument("--read", action="store_true")
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--value", default="")
    parser.add_argument("--value-file", default="")
    args = parser.parse_args()

    mode = (args.mode or "file").lower()
    if args.read:
        if mode == "running":
            value = _read_excel_running(args.path, args.sheet, args.cell)
        else:
            value = _read_excel_file(args.path, args.sheet, args.cell)
        print(value)
        return
    if args.write:
        value = args.value
        if args.value_file:
            value = Path(args.value_file).read_text(encoding="utf-8", errors="ignore")
        if mode == "running":
            _write_excel_running(args.path, args.sheet, args.cell, value)
        else:
            _write_excel_file(args.path, args.sheet, args.cell, value)
        return
    raise RuntimeError("Specify --read or --write.")


if __name__ == "__main__":
    main()
